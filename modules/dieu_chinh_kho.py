import os
import json
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright
import time
from colorama import Fore, Style, init
import sys
import shutil

# Khởi tạo colorama
init(autoreset=True)

class DieuChinhKho:
    def __init__(self):
        # Lấy đường dẫn thư mục chứa main.py hoặc file .exe
        if getattr(sys, 'frozen', False):
            # Chế độ đóng gói (PyInstaller)
            self.base_dir = os.path.dirname(sys.executable)
            self.resources_dir = os.path.join(self.base_dir, "resources")
        else:
            # Chế độ phát triển
            self.base_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
            self.resources_dir = self.base_dir  # Trong chế độ phát triển, sử dụng thư mục gốc

        # Đường dẫn đến Chromium
        self.chromium_path = os.path.join(self.resources_dir, "chromium-1169", "chrome-win", "chrome.exe")

        # Kiểm tra nếu Chromium tồn tại
        if not os.path.exists(self.chromium_path):
            raise FileNotFoundError(f"Không tìm thấy Chromium tại: {self.chromium_path}")

        self.parent_dir = os.path.join(self.base_dir, "dieu_chinh_kho")
        self.input_dir = os.path.join(self.parent_dir, "input")
        self.output_dir = os.path.join(self.parent_dir, "output")
        self.config_path = os.path.join(self.parent_dir, "config.json")
        self.template_path = os.path.join(self.input_dir, "template.xlsx")

        # Tạo thư mục nếu chưa tồn tại
        self.setup_directories()

        # Tạo file config mặc định nếu chưa tồn tại
        self.create_config_file()

    def setup_directories(self):
        for directory in [self.parent_dir, self.input_dir, self.output_dir]:
            if not os.path.exists(directory):
                os.makedirs(directory)
                print(Fore.GREEN + f"Đã tạo thư mục: {directory}")
            else:
                print(Fore.YELLOW + f"Thư mục đã tồn tại: {directory}")

    def create_config_file(self):
        if not os.path.exists(self.config_path):
            default_config = {
                "username": "admin",
                "password": "123456",
                "tool_name": "Tool All In One",
                "link": "https://example.com/login"
            }
            with open(self.config_path, "w", encoding="utf-8") as config_file:
                json.dump(default_config, config_file, indent=4)
            print(Fore.GREEN + f"Đã tạo file config: {self.config_path}")
        else:
            print(Fore.YELLOW + f"File config đã tồn tại: {self.config_path}")

    def read_config(self):
        if os.path.exists(self.config_path):
            with open(self.config_path, "r", encoding="utf-8") as config_file:
                return json.load(config_file)
        print(Fore.RED + "File config không tồn tại.")
        return None

    def create_excel_template(self):
        if os.path.exists(self.template_path):
            print(Fore.YELLOW + f"File template đã tồn tại: {self.template_path}")
            return

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Template"
        headers = ["Mã NPP", "Loại điều chỉnh", "Mã SP", "Loại kho", "Số lượng"]
        sheet.append(headers)
        sheet.freeze_panes = "A2"

        fill_color = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            cell = sheet[f"{col_letter}1"]
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = fill_color
            sheet.column_dimensions[col_letter].width = 25

        workbook.save(self.template_path)
        print(Fore.GREEN + f"Đã tạo file template: {self.template_path}")

    def process_and_create_files(self):
        # Kiểm tra nếu file template.xlsx không tồn tại
        if not os.path.exists(self.template_path):
            print(Fore.YELLOW + f"File template không tồn tại. Đang tạo file template tại: {self.template_path}")
            self.create_excel_template()

        # Đọc dữ liệu từ file template.xlsx
        workbook = load_workbook(self.template_path)
        sheet = workbook.active

        # Lấy dữ liệu từ file Excel (bỏ qua dòng tiêu đề)
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if all(row):  # Bỏ qua dòng trống
                data.append({
                    "Mã NPP": row[0],
                    "Loại điều chỉnh": row[1],
                    "Mã SP": row[2],
                    "Loại kho": row[3],
                    "Số lượng": row[4]
                })

        if not data:
            print(Fore.RED + "Không có dữ liệu hợp lệ trong file template.")
            return
        # Kiểm tra dữ liệu trong cột "Loại điều chỉnh"
        valid_adjustments = {"Nhập", "Xuất"}
        invalid_rows = [row for row in data if row["Loại điều chỉnh"] not in valid_adjustments]

        if invalid_rows:
            print(Fore.RED + "File Template không chuẩn. Vui lòng kiểm tra lại cột 'Loại điều chỉnh'.")
            print(Fore.RED + "Chỉ chấp nhận giá trị 'Nhập' hoặc 'Xuất'.")
            return
        
        # Xóa hết dữ liệu trong thư mục output trước khi tạo mới
        for file_name in os.listdir(self.output_dir):
            file_path = os.path.join(self.output_dir, file_name)
            try:
                if os.path.isfile(file_path):
                    os.unlink(file_path)
                    #print(Fore.GREEN + f"Đã xóa file: {file_path}")
            except Exception as e:
                print(Fore.RED + f"Lỗi khi xóa file {file_path}: {e}")

        # Tách dữ liệu theo Loại điều chỉnh
        data_xuat = [row for row in data if row["Loại điều chỉnh"] == "Xuất"]
        data_nhap = [row for row in data if row["Loại điều chỉnh"] == "Nhập"]

        # Xử lý từng loại điều chỉnh
        self._process_data(data_xuat, prefix="DCG")
        self._process_data(data_nhap, prefix="DCT")

    def _process_data(self, data, prefix):
        # Nhóm dữ liệu theo Mã NPP
        grouped_data = {}
        for row in data:
            key = row["Mã NPP"]
            if key not in grouped_data:
                grouped_data[key] = []
            grouped_data[key].append(row)

        # Tạo file Excel mới cho từng nhóm
        for ma_npp, rows in grouped_data.items():
            file_name = f"{prefix}_{ma_npp}.xlsx"
            file_path = os.path.join(self.output_dir, file_name)

            new_workbook = Workbook()
            new_sheet = new_workbook.active
            new_sheet.title = "Sheet1"  # Đặt tên sheet là "Sheet1"

            # Định dạng tiêu đề
            headers_row_1 = ["Loại điều chỉnh", "Mã đơn vị", "Loại đơn", "Số phiếu xuất"]
            headers_row_2 = ["Mã sản phẩm", "Loại kho", "Số lượng"]

            # Điền tiêu đề dòng 1
            for col_num, header in enumerate(headers_row_1, 1):
                col_letter = get_column_letter(col_num)
                cell = new_sheet[f"{col_letter}1"]
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
                new_sheet.column_dimensions[col_letter].width = 20

            # Điền tiêu đề dòng 3
            for col_num, header in enumerate(headers_row_2, 1):
                col_letter = get_column_letter(col_num)
                cell = new_sheet[f"{col_letter}3"]
                cell.value = header
                cell.font = Font(bold=True, color="FFA500")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.fill = PatternFill(start_color="228B22", end_color="228B22", fill_type="solid")
                new_sheet.column_dimensions[col_letter].width = 20

            # Điền dữ liệu vào ô A2 và B2
            new_sheet["A2"] = 1 if prefix == "DCG" else 0  # Loại điều chỉnh
            new_sheet["B2"] = ma_npp  # Mã đơn vị

            # Điền dữ liệu từ dòng 4
            for row_num, row in enumerate(rows, start=4):
                if "Mã SP" in row and "Loại kho" in row and "Số lượng" in row:
                    new_sheet[f"A{row_num}"] = row["Mã SP"]  # Mã sản phẩm
                    new_sheet[f"B{row_num}"] = row["Loại kho"]  # Loại kho
                    new_sheet[f"C{row_num}"] = row["Số lượng"]  # Số lượng
                else:
                    print(f"Dữ liệu thiếu khóa tại dòng {row_num}: {row}")

            # Lưu file
            new_workbook.save(file_path)
            print(Fore.GREEN + f"Đã tạo file: {file_path}")

    def import_dct_to_web(self):
        print(Fore.CYAN + "Bắt đầu import file DCT lên web...")
        self._import_files_to_web(file_prefix="DCT")
        print(Fore.GREEN + "Đã hoàn thành import file DCT lên web.")

    def import_dcg_to_web(self):
        print(Fore.CYAN + "Bắt đầu import file DCG lên web...")
        self._import_files_to_web(file_prefix="DCG")
        print(Fore.GREEN + "Đã hoàn thành import file DCG lên web.")

    def _import_files_to_web(self, file_prefix):
        # Đọc thông tin từ config
        config = self.read_config()
        if not config or "link" not in config:
            print("Link web không tồn tại trong config.")
            return

        link = config["link"]
        username = config["username"]
        password = config["password"]

        # Sử dụng Playwright để điều khiển trình duyệt
        with sync_playwright() as p:
            # Sử dụng Chromium từ thư mục `chromium-1169`
            browser = p.chromium.launch(
                headless=False,
                executable_path=self.chromium_path  # Đường dẫn đến trình duyệt Chromium
            )
            context = browser.new_context()
            page = context.new_page()

            try:
                # Điều hướng đến trang đăng nhập
                page.goto(link)

                # Nhập username và password
                page.fill('//*[@id="username"]', username)
                page.fill('//*[@id="password"]', password)
                page.click('//*[@id="fm1"]/fieldset/div[3]/input[3]')

                #print("Đăng nhập thành công.")
                time.sleep(1)  # Đợi một chút để trang tải xong

                # Lấy danh sách file trong thư mục output, chỉ lấy file có prefix tương ứng
                files = sorted(
                    [os.path.join(self.output_dir, f) for f in os.listdir(self.output_dir) if f.startswith(file_prefix) and f.endswith(".xlsx")],
                    key=lambda x: x
                )

                for file_path in files:
                    #print(f"Đang xử lý file: {file_path}")
                    page.goto(link)
                    page.click('//*[@id="1257"]/a')
                    page.click('//*[@id="child_1264"]')
                    page.click('//*[@id="child_Child_1267"]')

                    # Điều hướng đến form nhập file
                    page.click('//*[@id="btnInputFile_"]')

                    # Đợi modal hiển thị
                    page.wait_for_selector('//*[@id="excelFileStockUpdate"]')

                    # Chọn file bằng nút "Browse"
                    canonical_path = os.path.abspath(file_path)
                    page.set_input_files('//*[@id="excelFileStockUpdate"]', canonical_path)
                    #print(f"Đã chọn file: {canonical_path}")

                    # Nhấn nút "Nhập Excel"
                    page.click('//*[@id="grpImport_dlg"]/div/div[4]/button[1]')
                    #print("Đã nhấn nút Nhập Excel.")

                    # Đợi popup xác nhận hiển thị
                    page.wait_for_selector("//div[contains(@class, 'messager-window')]")
                    time.sleep(2)

                    # Sử dụng XPath để nhấn nút "Đồng ý"
                    page.locator('//html/body/div[35]/div[2]/div[4]/a[1]/span/span').click()
                    #print("Đã nhấn nút Đồng ý.")

                    time.sleep(3)
                    # Đợi thông báo kết quả
                    result_message = page.inner_text("//p[@id='errExcelMsgStockUpdate']")
                    print(f"- Kết quả: {os.path.basename(file_path)} - {result_message}")

                    # Làm mới trang
                    page.goto(link)
                    #print("Đã làm mới trang.")

            except Exception as e:
                print(f"Lỗi khi thực hiện thao tác: {e}")
                import traceback
                traceback.print_exc()
            finally:
                browser.close()
